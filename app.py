import streamlit as st
import pandas as pd
import fitz  # PyMuPDF
import re
import openpyxl
from datetime import datetime

st.set_page_config(page_title="Analisador Pós-Solar", layout="centered")
st.title("☀️ Analisador de Geração e Consumo - Pós Energia Solar")

fatura = st.file_uploader("📄 Enviar fatura (PDF):", type=["pdf"])
geracoes = st.file_uploader("📊 Enviar dois relatórios de geração (XLS):", type=["xls", "xlsx"], accept_multiple_files=True)

def extrair_texto_pdf(fatura):
    texto = ""
    with fitz.open(stream=fatura.read(), filetype="pdf") as doc:
        for page in doc:
            texto += page.get_text()
    return texto

def extrair_dados_fatura(texto):
    data_match = re.search(r"(\d{2}/\d{2}/\d{4})\s+(\d{2}/\d{2}/\d{4})", texto)
    consumo_match = re.search(r"ENERGIA ELET CONSUMO.*?(\d{1,5})", texto)
    injetada_match = re.search(r"ENERGIA INJETADA.*?(\d{1,5})", texto)
    creditos_match = re.search(r"Saldo Acumulado.*?(\d{1,5})", texto)

    data_inicio = datetime.strptime(data_match.group(1), "%d/%m/%Y") if data_match else None
    data_fim = datetime.strptime(data_match.group(2), "%d/%m/%Y") if data_match else None
    consumo = int(consumo_match.group(1)) if consumo_match else 0
    injetada = int(injetada_match.group(1)) if injetada_match else 0
    creditos = int(creditos_match.group(1)) if creditos_match else 0

    return consumo, injetada, creditos, data_inicio, data_fim

def calcular_geracao(arquivos, data_inicio, data_fim):
    total = 0.0
    for arquivo in arquivos:
        wb = openpyxl.load_workbook(arquivo, data_only=True)
        for nome_aba in wb.sheetnames:
            aba = wb[nome_aba]
            headers = [cell.value for cell in aba[1]]
            if "Time" in headers and "Yield(kWh)" in headers:
                idx_data = headers.index("Time") + 1
                idx_kwh = headers.index("Yield(kWh)") + 1
                for row in aba.iter_rows(min_row=2):
                    data_val = row[idx_data - 1].value
                    kwh_val = row[idx_kwh - 1].value
                    if isinstance(data_val, datetime) and data_inicio <= data_val <= data_fim:
                        try:
                            total += float(kwh_val)
                        except:
                            pass
    return total

if fatura and geracoes and len(geracoes) == 2:
    st.markdown(f"### 📂 Análise: `{fatura.name}` + 2 arquivos de geração")
    texto = extrair_texto_pdf(fatura)
    consumo, injetada, creditos, data_inicio, data_fim = extrair_dados_fatura(texto)
    
    st.text_area("📑 Texto extraído da fatura:", texto, height=250)

    if not data_inicio or not data_fim:
        st.error("⚠️ Não foi possível identificar o período da fatura automaticamente.")
    else:
        gerado = calcular_geracao(geracoes, data_inicio, data_fim)
        total_utilizado = consumo + injetada
        eficiencia = (gerado / total_utilizado * 100) if total_utilizado > 0 else 0
        desempenho = (injetada / gerado * 100) if gerado > 0 else 0

        st.subheader("📊 Resultados")
        st.write(f"📆 Período informado: **{data_inicio.strftime('%d/%m/%Y')} a {data_fim.strftime('%d/%m/%Y')}**")
        st.write(f"🌞 Geração total no período: **{gerado:.2f} kWh**")
        st.write(f"⚡ Consumo instantâneo (da rede): **{consumo} kWh**")
        st.write(f"🔁 Energia injetada na rede: **{injetada} kWh**")
        st.write(f"💳 Créditos acumulados: **{creditos} kWh**")
        st.write(f"📐 Eficiência de uso da geração: **{eficiencia:.1f}%**")
        st.write(f"🎯 Desempenho da geração vs. meta: **{desempenho:.1f}%**")

        st.subheader("💡 Sugestões")
        if gerado == 0:
            st.warning("⚠️ Geração abaixo do esperado: verificar sombreamentos ou falhas no sistema.")
        if desempenho < 70:
            st.info("📉 Baixo desempenho de geração: possível problema de dimensionamento.")
        if eficiencia < 70:
            st.info("🔌 Baixa eficiência de uso: pode haver subutilização da geração.")
        if injetada > gerado * 0.7:
            st.info("🔥 Alta injeção na rede: consumo local está baixo, considerar redimensionar.")
else:
    st.info("📌 Envie uma fatura e exatamente **2 arquivos de geração** para a análise.")
